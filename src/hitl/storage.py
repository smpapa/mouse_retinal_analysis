"""Read/write `oct_results.xlsx` for the HITL editor.

Layout:
  - The xlsx has one `summary` sheet plus one detail sheet per image.
  - Each detail sheet's columns are the per-x measurements produced by
    the batch pipeline. We treat the auto-detection columns as immutable
    and store user edits in parallel `<name>_corrected` columns.

Workbook -> memory:
  load_workbook(path) -> Workbook
    .images: dict[stem, ImageRecord]

Memory -> workbook (separate task)
"""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import numpy as np
import pandas as pd

from src.hitl.boundary_model import (BOUNDARY_NAMES, ERASED_MARKER,
                                       ERASED_THRESHOLD)


# Re-exported for backwards compatibility — older callers import AUTO_COLS
# directly from `storage`. Both names refer to the same canonical tuple.
AUTO_COLS = BOUNDARY_NAMES

# String literal written to xlsx to represent an erased (explicit-NaN) cell.
ERASED_CELL_TEXT = "ERASED"


def _is_set_cell(v) -> bool:
    """True if a corrected-column cell holds a user edit (number or ERASED)."""
    if isinstance(v, str):
        return v == ERASED_CELL_TEXT
    try:
        return not np.isnan(float(v))
    except (TypeError, ValueError):
        return False


@dataclass
class ImageRecord:
    stem: str                  # e.g. "21_OS_4H"
    filename: str              # e.g. "21_OS_4H.tif"
    width: int                 # B-scan column count
    auto: dict[str, np.ndarray] = field(default_factory=dict)
    corrected: dict[str, np.ndarray] = field(default_factory=dict)


@dataclass
class Workbook:
    path: Path
    summary: pd.DataFrame
    images: dict[str, ImageRecord] = field(default_factory=dict)


def load_workbook(path: str | Path) -> Workbook:
    p = Path(path)
    xls = pd.ExcelFile(p)
    summary = pd.read_excel(xls, sheet_name="summary")
    wb = Workbook(path=p, summary=summary)

    # Sheet names may be truncated to 28 chars (with `_2`, `_3` for duplicates),
    # so we cannot derive the original filename/stem from the sheet name.
    # The batch writes summary rows and detail sheets in matching order, so we
    # zip the detail sheets against the `filename` column from the summary.
    if "filename" in summary.columns:
        summary_filenames = summary["filename"].tolist()
    else:
        summary_filenames = []
    fn_iter = iter(summary_filenames)

    for sheet in xls.sheet_names:
        if sheet == "summary":
            continue
        if sheet == "corrected_summary":
            continue
        df = pd.read_excel(xls, sheet_name=sheet)
        if "x_local" not in df.columns:
            continue
        try:
            filename = next(fn_iter)
        except StopIteration:
            # No corresponding summary row (e.g. unrelated future sheet).
            continue
        width = len(df)
        stem = Path(str(filename)).stem
        rec = ImageRecord(stem=stem, filename=str(filename), width=width)
        for name in AUTO_COLS:
            if name in df.columns:
                rec.auto[name] = pd.to_numeric(
                    df[name], errors="coerce"
                ).to_numpy(dtype=float)
            else:
                rec.auto[name] = np.full(width, np.nan, dtype=float)
            corr_col = f"{name}_corrected"
            if corr_col in df.columns:
                # Map literal "ERASED" cells back to ERASED_MARKER so the
                # sentinel survives save->reload (pd.to_numeric would
                # otherwise silently coerce it to NaN, indistinguishable
                # from "untouched").
                raw = df[corr_col]
                out = np.full(width, np.nan, dtype=float)
                nums = pd.to_numeric(raw, errors="coerce").to_numpy(dtype=float)
                is_erased = raw.astype(str).str.upper() == ERASED_CELL_TEXT
                out[:] = nums
                out[is_erased.to_numpy()] = ERASED_MARKER
                rec.corrected[name] = out
            else:
                rec.corrected[name] = np.full(width, np.nan, dtype=float)
        wb.images[stem] = rec
    return wb


@dataclass
class CorrectedSnapshot:
    stem: str
    corrected: dict[str, np.ndarray]
    timestamp: datetime


def _read_column(ws, col_idx: int, n_rows: int) -> np.ndarray:
    """Read a numeric column from openpyxl worksheet (1-based col_idx,
    skipping the header row). Non-numeric / blank cells become NaN."""
    out = np.full(n_rows, np.nan, dtype=float)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=n_rows + 1,
                                          min_col=col_idx, max_col=col_idx,
                                          values_only=True)):
        v = row[0]
        if isinstance(v, (int, float)) and not (
                isinstance(v, float) and np.isnan(v)):
            out[i] = float(v)
    return out


def _read_corrected_column(ws, col_idx: int, n_rows: int) -> np.ndarray:
    """Like _read_column, but maps the literal "ERASED" string to NaN
    and returns the parsed numeric column for thickness recompute."""
    out = np.full(n_rows, np.nan, dtype=float)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=n_rows + 1,
                                          min_col=col_idx, max_col=col_idx,
                                          values_only=True)):
        v = row[0]
        if v == ERASED_CELL_TEXT:
            out[i] = np.nan
        elif isinstance(v, (int, float)) and not (
                isinstance(v, float) and np.isnan(v)):
            out[i] = float(v)
    return out


def apply_corrections_inplace(wb,
                               snapshots: list[CorrectedSnapshot],
                               scale_um_per_px_y: float) -> None:
    """Mutate an in-memory openpyxl Workbook with the given snapshots.

    Splits the work that ``save_corrections`` does so the caller can
    cache the workbook between saves (avoiding the ~10 s reload cost on
    a 96-sheet xlsx) and run the disk flush on a background thread.
    """
    # Build stem -> sheet_name map by walking the summary sheet's
    # `filename` column in order. The batch pipeline writes detail sheets
    # in matching order, so we zip them together.
    stem_to_sheet: dict[str, str] = {}
    summary_ws = wb["summary"] if "summary" in wb.sheetnames else None
    if summary_ws is not None:
        headers = [c.value for c in summary_ws[1]]
        if "filename" in headers:
            fn_col = headers.index("filename") + 1
            data_sheets = [s for s in wb.sheetnames
                           if s not in ("summary", "corrected_summary")]
            for row_idx, sheet_name in enumerate(data_sheets, start=2):
                cell_value = summary_ws.cell(row=row_idx, column=fn_col).value
                if cell_value:
                    stem = Path(str(cell_value)).stem
                    stem_to_sheet[stem] = sheet_name

    summary_rows: list[dict] = []

    for snap in snapshots:
        sheet_name = stem_to_sheet.get(snap.stem)
        if sheet_name is None or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        # Locate or create columns by header.
        headers = [c.value for c in ws[1]]
        col_idx: dict[str, int] = {
            h: i + 1 for i, h in enumerate(headers) if h is not None
        }
        n_rows = ws.max_row - 1  # excludes header

        def ensure_col(name: str) -> int:
            if name in col_idx:
                return col_idx[name]
            new_idx = len(col_idx) + 1
            ws.cell(row=1, column=new_idx, value=name)
            col_idx[name] = new_idx
            return new_idx

        # Write each boundary's corrected column.
        any_corrected: dict[str, bool] = {}
        for name, arr in snap.corrected.items():
            c = ensure_col(f"{name}_corrected")
            for i, v in enumerate(arr.tolist()):
                if v < ERASED_THRESHOLD:
                    cell_value = ERASED_CELL_TEXT
                elif np.isnan(v):
                    cell_value = None
                else:
                    cell_value = float(v)
                ws.cell(row=i + 2, column=c, value=cell_value)
            erased_mask = arr < ERASED_THRESHOLD
            finite_mask = (~np.isnan(arr)) & (~erased_mask)
            any_corrected[name] = bool(erased_mask.any() or finite_mask.any())

        # Recompute thicknesses using effective boundary values
        # (auto where no correction, corrected otherwise).
        eff: dict[str, np.ndarray] = {}
        for name in AUTO_COLS:
            if name in col_idx:
                auto = _read_column(ws, col_idx[name], n_rows)
            else:
                auto = np.full(n_rows, np.nan, dtype=float)
            corr_col_name = f"{name}_corrected"
            if corr_col_name in col_idx:
                corr = _read_corrected_column(ws, col_idx[corr_col_name],
                                                n_rows)
                # The corrected-column read just gives finite values where
                # set OR explicit NaN where erased; auto is the fallback.
                # Build effective: corr if cell text was "ERASED" -> NaN,
                # else corr if finite, else auto.
                # Easier: re-read with a more explicit pass.
                erased = np.zeros(n_rows, dtype=bool)
                for i, row in enumerate(ws.iter_rows(
                        min_row=2, max_row=n_rows + 1,
                        min_col=col_idx[corr_col_name],
                        max_col=col_idx[corr_col_name],
                        values_only=True)):
                    erased[i] = (row[0] == ERASED_CELL_TEXT)
                out = auto.copy()
                # finite corrected → override
                set_mask = ~np.isnan(corr) & ~erased
                out[set_mask] = corr[set_mask]
                out[erased] = np.nan
                eff[name] = out
            else:
                eff[name] = auto

        # Write thickness columns.
        total = (eff["BM_y"] - eff["TOP_y"]) * scale_um_per_px_y
        outer = (eff["BM_y"] - eff["ONL_y"]) * scale_um_per_px_y
        det = (eff["DET_bottom_y"] - eff["DET_top_y"]) * scale_um_per_px_y
        for col_name, vals in (
            ("total_thickness_um_corrected", total),
            ("outer_thickness_um_corrected", outer),
            ("detachment_thickness_um_corrected", det),
        ):
            c = ensure_col(col_name)
            for i, v in enumerate(vals.tolist()):
                ws.cell(row=i + 2, column=c,
                        value=(None if np.isnan(v) else float(v)))

        # corrected_by_user flag per row.
        c_flag = ensure_col("corrected_by_user")
        any_per_col = np.zeros(n_rows, dtype=bool)
        for name in AUTO_COLS:
            corr_col_name = f"{name}_corrected"
            if corr_col_name not in col_idx:
                continue
            cc = col_idx[corr_col_name]
            for i, row in enumerate(ws.iter_rows(
                    min_row=2, max_row=n_rows + 1,
                    min_col=cc, max_col=cc, values_only=True)):
                if _is_set_cell(row[0]):
                    any_per_col[i] = True
        for i, v in enumerate(any_per_col.tolist()):
            ws.cell(row=i + 2, column=c_flag, value=bool(v))

        # Build the corrected_summary row.
        n_corr = int(any_per_col.sum())
        if "total_thickness_um" in col_idx:
            auto_total = _read_column(ws, col_idx["total_thickness_um"], n_rows)
            mean_total = float(np.nanmean(auto_total)) \
                if not np.all(np.isnan(auto_total)) else float("nan")
        else:
            mean_total = float("nan")
        mean_total_corr = float(np.nanmean(total)) \
            if not np.all(np.isnan(total)) else float("nan")
        summary_rows.append({
            "filename": f"{snap.stem}.tif",
            "n_corrected_cols": n_corr,
            "corrected_TOP": bool(any_corrected.get("TOP_y", False)),
            "corrected_ONL": bool(any_corrected.get("ONL_y", False)),
            "corrected_BM": bool(any_corrected.get("BM_y", False)),
            "corrected_DET": bool(any_corrected.get("DET_top_y", False)
                                  or any_corrected.get("DET_bottom_y", False)),
            "mean_total_um": mean_total,
            "mean_total_um_corrected": mean_total_corr,
            "edit_timestamp": snap.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
        })

    # corrected_summary: append/replace rows for the saved stems.
    if summary_rows:
        cs_headers = ["filename", "n_corrected_cols",
                      "corrected_TOP", "corrected_ONL",
                      "corrected_BM", "corrected_DET",
                      "mean_total_um", "mean_total_um_corrected",
                      "edit_timestamp"]
        if "corrected_summary" in wb.sheetnames:
            cs = wb["corrected_summary"]
            # Build filename -> row map (1-based, includes header row).
            existing_headers = [c.value for c in cs[1]]
            try:
                fn_idx = existing_headers.index("filename") + 1
            except ValueError:
                fn_idx = 1
            existing_rows: dict[str, int] = {}
            for r in range(2, cs.max_row + 1):
                v = cs.cell(row=r, column=fn_idx).value
                if v:
                    existing_rows[str(v)] = r
            # Map header → 1-based col index in the existing sheet.
            ex_col_idx = {h: i + 1 for i, h in enumerate(existing_headers)
                          if h is not None}
            # Ensure all our headers exist.
            for hname in cs_headers:
                if hname not in ex_col_idx:
                    new_idx = len(ex_col_idx) + 1
                    cs.cell(row=1, column=new_idx, value=hname)
                    ex_col_idx[hname] = new_idx
            for srow in summary_rows:
                fn = srow["filename"]
                target_row = existing_rows.get(fn, cs.max_row + 1)
                for k, v in srow.items():
                    cs.cell(row=target_row, column=ex_col_idx[k], value=v)
        else:
            cs = wb.create_sheet("corrected_summary")
            for j, h in enumerate(cs_headers, start=1):
                cs.cell(row=1, column=j, value=h)
            for ri, srow in enumerate(summary_rows, start=2):
                for j, h in enumerate(cs_headers, start=1):
                    cs.cell(row=ri, column=j, value=srow[h])



def save_workbook_atomic(wb, path: str | Path) -> None:
    """Serialize an in-memory openpyxl Workbook to disk atomically.

    Writes to ``<path>.tmp`` first, then ``os.replace`` over the target
    so a crash mid-write cannot corrupt the user's xlsx.
    """
    p = Path(path)
    tmp = p.with_suffix(p.suffix + ".tmp")
    wb.save(str(tmp))
    os.replace(str(tmp), str(p))


def save_corrections(path: str | Path,
                     snapshots: list[CorrectedSnapshot],
                     scale_um_per_px_y: float,
                     wb=None) -> None:
    """Apply ``snapshots`` to the workbook and save to disk atomically.

    If ``wb`` is None, the workbook is loaded fresh from ``path``; this
    is the path tests and one-off CLI use take. The HITL editor caches
    the workbook between saves and passes it via ``wb=`` to skip the
    ~10 s reload cost on a 96-sheet xlsx.
    """
    from openpyxl import load_workbook as _openpyxl_load
    own_wb = (wb is None)
    if own_wb:
        wb = _openpyxl_load(str(path))
    try:
        apply_corrections_inplace(wb, snapshots, scale_um_per_px_y)
        save_workbook_atomic(wb, path)
    finally:
        if own_wb:
            wb.close()
