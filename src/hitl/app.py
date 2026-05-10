"""Main window: sidebar + canvas + storage glue."""
from __future__ import annotations

import sys
import traceback
from datetime import datetime
from pathlib import Path

import numpy as np
from PySide6.QtCore import QObject, Qt, QThread, Signal
from PySide6.QtGui import QAction, QActionGroup, QKeySequence
from PySide6.QtWidgets import (QApplication, QDockWidget, QFileDialog,
                                QLabel, QMainWindow, QMessageBox,
                                QProgressDialog, QStatusBar, QToolBar,
                                QVBoxLayout, QWidget)

# Make sibling analysis modules importable.
SRC = Path(__file__).resolve().parents[1]
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from io_utils import load_oct                                       # noqa: E402

from .batch_runner import BatchWorker                                # noqa: E402
from .boundary_model import (BOUNDARY_NAMES, BoundaryEditor,
                              ERASED_THRESHOLD)                      # noqa: E402
from .boundary_toggle import BoundaryToggleBar                      # noqa: E402
from .canvas import EditMode, OverlayCanvas                         # noqa: E402
from .overlay_render import render_corrected_overlay                # noqa: E402
from .sidebar import FileEntry, FileListView                        # noqa: E402
from .storage import (CorrectedSnapshot, Workbook,
                      load_workbook, save_corrections)              # noqa: E402


class _SaveWorker(QObject):
    """Background-thread worker that flushes a cached openpyxl Workbook
    to disk via the storage helpers."""
    finished = Signal()
    error = Signal(str)

    def __init__(self, wb, path):
        super().__init__()
        self.wb = wb
        self.path = path

    def run(self):
        try:
            from .storage import save_workbook_atomic
            save_workbook_atomic(self.wb, self.path)
            self.finished.emit()
        except Exception as e:
            self.error.emit(f"{e}\n\n{traceback.format_exc()}")


class MainWindow(QMainWindow):
    def __init__(self, workbook_path, image_dir):
        super().__init__()
        self.workbook_path = Path(workbook_path)
        self.image_dir = Path(image_dir)
        self._editors: dict[str, BoundaryEditor] = {}
        self._wb: Workbook | None = None
        self._current_stem: str | None = None
        self._scale_y = 3.87  # default; overridden if summary has it.
        # When True, _on_sidebar_image_selected skips the unsaved-changes
        # prompt — used for programmatic reverts after the user cancels.
        self._suppress_dirty_check: bool = False
        # Batch-run worker state. Held as instance attrs so the QThread
        # outlives the method scope; cleared after batch completes.
        self._batch_thread: QThread | None = None
        self._batch_worker: BatchWorker | None = None
        self._progress_dialog: QProgressDialog | None = None
        # Cached openpyxl workbook (loaded lazily on first save) so we
        # don't pay the ~10 s reload cost every Ctrl+S. Invalidated when
        # the user switches data folders.
        self._openpyxl_wb = None
        # Background save state.
        self._save_thread: QThread | None = None
        self._save_worker = None
        self._save_in_flight: bool = False
        self._pending_snapshots: list = []  # of CorrectedSnapshot
        self._currently_saving_stems: list[str] = []

        self._build_ui()
        # If a real workbook exists, load it. Otherwise leave the UI empty
        # so the user can pick a folder via File > Open Data Folder.
        if self.workbook_path.exists():
            self._reload_workbook()

    def _build_ui(self) -> None:
        self.setWindowTitle("OCT HITL Editor")
        self.canvas = OverlayCanvas()
        self.canvas.edit_finished.connect(self._refresh_status)
        self.setCentralWidget(self.canvas)

        self.sidebar = FileListView()
        # Route through the dirty-check handler instead of select_image
        # directly so we can prompt before discarding unsaved edits.
        self.sidebar.image_selected.connect(self._on_sidebar_image_selected)

        self.boundary_toggle = BoundaryToggleBar()
        self.boundary_toggle.visibility_changed.connect(
            self.canvas.set_boundary_visible
        )

        container = QWidget()
        container_layout = QVBoxLayout(container)
        container_layout.setContentsMargins(0, 0, 0, 0)
        container_layout.addWidget(self.sidebar, stretch=1)
        container_layout.addWidget(self.boundary_toggle)

        dock = QDockWidget("Files", self)
        dock.setWidget(container)
        dock.setAllowedAreas(Qt.LeftDockWidgetArea)
        self.addDockWidget(Qt.LeftDockWidgetArea, dock)

        toolbar = QToolBar("Edit", self)
        self.addToolBar(toolbar)

        self.act_drag = QAction("Drag", self)
        self.act_drag.setCheckable(True)
        self.act_drag.setChecked(True)
        self.act_drag.triggered.connect(
            lambda: self.canvas.set_mode(EditMode.DRAG))
        self.act_erase = QAction("Erase", self)
        self.act_erase.setCheckable(True)
        self.act_erase.triggered.connect(
            lambda: self.canvas.set_mode(EditMode.ERASE))
        # Mutually exclusive selection between drag and erase.
        mode_group = QActionGroup(self)
        mode_group.setExclusive(True)
        mode_group.addAction(self.act_drag)
        mode_group.addAction(self.act_erase)
        toolbar.addAction(self.act_drag)
        toolbar.addAction(self.act_erase)

        toolbar.addSeparator()
        self.act_undo = QAction("Undo", self)
        self.act_undo.setShortcut(QKeySequence.Undo)
        self.act_undo.triggered.connect(self._undo)
        toolbar.addAction(self.act_undo)

        self.act_save = QAction("Save", self)
        self.act_save.setShortcut(QKeySequence.Save)
        self.act_save.triggered.connect(self.save_current_image)
        toolbar.addAction(self.act_save)

        self._setup_shortcuts()
        self._build_menu_bar()
        self.setStatusBar(QStatusBar(self))
        # Persistent hover label (rightmost in the status bar). Updated
        # whenever the canvas reports the column under the cursor; the
        # transient showMessage() text on the left of the status bar is
        # untouched.
        self._hover_label = QLabel("")
        self._hover_label.setStyleSheet("padding-right: 8px;")
        self.statusBar().addPermanentWidget(self._hover_label)
        self.canvas.hovered.connect(self._on_canvas_hover)

    def _build_menu_bar(self) -> None:
        """File / Tools menus driving folder selection + auto analysis."""
        menubar = self.menuBar()

        file_menu = menubar.addMenu("&File")
        self.act_open = QAction("&Open Data Folder...", self)
        self.act_open.setShortcut(QKeySequence.Open)
        self.act_open.triggered.connect(self._open_data_folder)
        file_menu.addAction(self.act_open)
        file_menu.addSeparator()
        # Reuse the toolbar's Save action so its shortcut and label match.
        file_menu.addAction(self.act_save)

        tools_menu = menubar.addMenu("&Tools")
        self.act_run_batch = QAction("&Run Auto Analysis...", self)
        self.act_run_batch.triggered.connect(self._run_auto_analysis)
        tools_menu.addAction(self.act_run_batch)

    def _setup_shortcuts(self) -> None:
        # Boundary picker shortcuts: 1..5 select active boundary.
        keys = ["1", "2", "3", "4", "5"]
        for k, name in zip(keys, BOUNDARY_NAMES):
            act = QAction(self)
            act.setShortcut(QKeySequence(k))
            act.triggered.connect(lambda _=False, n=name:
                                  self.canvas.set_active_boundary(n))
            self.addAction(act)

        # Mode shortcuts: D = drag, E = erase. Trigger the action so the
        # checked state in the QActionGroup updates alongside the mode.
        act_d = QAction(self)
        act_d.setShortcut(QKeySequence("D"))
        act_d.triggered.connect(self._activate_drag_mode)
        self.addAction(act_d)

        act_e = QAction(self)
        act_e.setShortcut(QKeySequence("E"))
        act_e.triggered.connect(self._activate_erase_mode)
        self.addAction(act_e)

        # Image navigation: Left/Right arrows step through the sidebar.
        act_prev = QAction(self)
        act_prev.setShortcut(QKeySequence(Qt.Key_Left))
        act_prev.triggered.connect(self._prev_image)
        self.addAction(act_prev)

        act_next = QAction(self)
        act_next.setShortcut(QKeySequence(Qt.Key_Right))
        act_next.triggered.connect(self._next_image)
        self.addAction(act_next)

    def _activate_drag_mode(self) -> None:
        self.canvas.set_mode(EditMode.DRAG)
        self.act_drag.setChecked(True)

    def _activate_erase_mode(self) -> None:
        self.canvas.set_mode(EditMode.ERASE)
        self.act_erase.setChecked(True)

    def _reload_workbook(self) -> None:
        self._wb = load_workbook(self.workbook_path)
        # Pull scale_um_per_px_y from summary if present.
        if "scale_um_per_px_y" in self._wb.summary.columns:
            try:
                self._scale_y = float(self._wb.summary["scale_um_per_px_y"].iloc[0])
            except (TypeError, ValueError):
                pass
        entries = []
        for stem, rec in self._wb.images.items():
            has_corr = any(np.any(~np.isnan(arr))
                           for arr in rec.corrected.values())
            entries.append(FileEntry(stem=stem,
                                     filename=rec.filename,
                                     has_corrections=has_corr))
        entries.sort(key=lambda e: e.stem)
        self.sidebar.set_entries(entries)

    def _on_sidebar_image_selected(self, stem: str) -> None:
        """Sidebar selection handler with unsaved-changes confirmation.

        If the previously-selected image has dirty edits, prompt the user
        to save / discard / cancel before switching. On cancel, revert
        the sidebar selection without re-prompting.
        """
        if self._suppress_dirty_check:
            self.select_image(stem)
            return
        prev = self._current_stem
        if (prev is not None
                and prev != stem
                and self._editors.get(prev) is not None
                and self._editors[prev].dirty):
            current_filename = (self._wb.images[prev].filename
                                if self._wb is not None else prev)
            reply = QMessageBox.question(
                self, "Unsaved changes",
                f"Save changes to {current_filename} before switching?",
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                QMessageBox.Save,
            )
            if reply == QMessageBox.Cancel:
                # Revert sidebar selection back to the previous stem
                # without re-firing the dirty prompt.
                self._suppress_dirty_check = True
                try:
                    if prev in self.sidebar._stem_for_row:
                        row = self.sidebar._stem_for_row.index(prev)
                        self.sidebar.setCurrentRow(row)
                finally:
                    self._suppress_dirty_check = False
                return
            if reply == QMessageBox.Save:
                self.save_current_image()
            # Discard: fall through and load the new image.
        self.select_image(stem)

    def select_image(self, stem: str) -> None:
        if self._wb is None or stem not in self._wb.images:
            return
        rec = self._wb.images[stem]
        img_path = self.image_dir / rec.filename
        # Load the TIFF first so a missing/corrupt file doesn't leave the
        # window pointing at a broken image. _current_stem is only updated
        # after the load succeeds.
        try:
            img = load_oct(img_path)
        except Exception as e:
            self.statusBar().showMessage(
                f"Could not load {rec.filename}: {e}"
            )
            return
        self._current_stem = stem
        if stem not in self._editors:
            self._editors[stem] = BoundaryEditor(
                width=rec.width,
                auto=rec.auto,
                corrected={k: v.copy() for k, v in rec.corrected.items()},
            )
        self.canvas.set_image(img.rgb, offset_x=img.layout.left_x)
        self.canvas.set_panel_geometry(
            left_x=img.layout.left_x,
            right_x=img.layout.right_x,
            top_y=img.layout.top_y,
            bot_y=img.layout.bot_y,
            center_x=img.layout.center_x,
        )
        self.canvas.set_editor(self._editors[stem])
        self.canvas.set_active_boundary("TOP_y")
        self._refresh_status()

    def save_current_image(self) -> None:
        """Queue the current image's edits for an asynchronous flush.

        UI feedback is immediate: sidebar gains the ✓ marker and the PNG
        is rendered up-front. The slow xlsx write happens on a background
        thread so the user can move to the next image right away.
        """
        if self._current_stem is None:
            return
        editor = self._editors[self._current_stem]
        snap = CorrectedSnapshot(
            stem=self._current_stem,
            corrected={k: editor.corrected[k].copy()
                       for k in editor.corrected},
            timestamp=datetime.now(),
        )
        self._pending_snapshots.append(snap)
        # Optimistic UI: sidebar ✓ + render PNG immediately. mark_clean
        # is deferred until the worker confirms the disk write so a
        # failure leaves the editor dirty for retry.
        self.sidebar.mark_corrected(self._current_stem, True)
        rec = self._wb.images[self._current_stem]
        try:
            img_path = self.image_dir / rec.filename
            out_path = (self.workbook_path.parent /
                        f"{self._current_stem}_overlay_corrected.png")
            boundaries = {k: editor.effective(k) for k in editor.auto}
            render_corrected_overlay(img_path, boundaries, out_path)
        except Exception as e:
            self.statusBar().showMessage(
                f"Save queued for {rec.filename} (overlay PNG failed: {e})")
        else:
            self.statusBar().showMessage(f"Saving {rec.filename}...")
        # Kick off the disk flush if no save is currently in flight.
        if not self._save_in_flight:
            self._kick_save_cycle()

    def _kick_save_cycle(self) -> None:
        """Apply queued snapshots to the cached workbook and spawn a
        background thread to flush it to disk. Caller must guarantee
        no save is currently in flight."""
        if not self._pending_snapshots:
            return
        from openpyxl import load_workbook as _openpyxl_load
        from .storage import apply_corrections_inplace, save_workbook_atomic

        snaps = self._pending_snapshots[:]
        self._pending_snapshots.clear()
        # Lazy-load the openpyxl workbook on first save; reused after.
        if self._openpyxl_wb is None:
            try:
                self._openpyxl_wb = _openpyxl_load(str(self.workbook_path))
            except Exception as e:
                QMessageBox.critical(self, "Save failed",
                                      f"Could not open workbook: {e}")
                return
        # Apply mutations to the in-memory workbook on the main thread
        # (the worker only reads it). This is fast (~50 ms per snap).
        try:
            apply_corrections_inplace(
                self._openpyxl_wb, snaps, self._scale_y,
            )
        except Exception as e:
            QMessageBox.critical(self, "Save failed",
                                  f"Could not apply corrections: {e}")
            return
        # Spawn the background flush.
        self._save_in_flight = True
        self._currently_saving_stems = [s.stem for s in snaps]
        self._save_thread = QThread(self)
        self._save_worker = _SaveWorker(self._openpyxl_wb,
                                          self.workbook_path)
        self._save_worker.moveToThread(self._save_thread)
        self._save_thread.started.connect(self._save_worker.run)
        self._save_worker.finished.connect(self._on_save_finished)
        self._save_worker.error.connect(self._on_save_error)
        self._save_thread.start()

    def _on_save_finished(self) -> None:
        # Mark editors clean now that the disk write completed.
        for stem in self._currently_saving_stems:
            ed = self._editors.get(stem)
            if ed is not None:
                ed.mark_clean()
        if self._save_thread is not None:
            self._save_thread.quit()
            self._save_thread.wait()
        self._save_thread = None
        self._save_worker = None
        self._save_in_flight = False
        stems = self._currently_saving_stems
        self._currently_saving_stems = []
        if stems:
            last = stems[-1]
            rec = self._wb.images.get(last) if self._wb else None
            label = rec.filename if rec is not None else last
            self.statusBar().showMessage(f"Saved {label}")
        self._refresh_status()
        # Process anything queued during the just-finished flush.
        if self._pending_snapshots:
            self._kick_save_cycle()

    def _on_save_error(self, msg: str) -> None:
        # Clean up any leftover .tmp from a partial write.
        tmp = self.workbook_path.with_suffix(
            self.workbook_path.suffix + ".tmp")
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        if self._save_thread is not None:
            self._save_thread.quit()
            self._save_thread.wait()
        self._save_thread = None
        self._save_worker = None
        self._save_in_flight = False
        self._currently_saving_stems = []
        self.statusBar().showMessage("Save failed")
        QMessageBox.critical(
            self, "Save failed",
            f"Could not write {self.workbook_path.name}.\n\n"
            f"If Excel has the file open, close it and try again.\n\n"
            f"Details: {msg}",
        )

    def _undo(self) -> None:
        if self._current_stem is None:
            return
        editor = self._editors[self._current_stem]
        editor.undo()
        self.canvas.refresh()
        self._refresh_status()

    # ------------------------------------------------------ data folder

    def _open_data_folder(self) -> None:
        """Prompt the user for a TIFF folder and switch to it.

        If an `output/oct_results.xlsx` already exists in that folder we
        load it directly; otherwise offer to run auto analysis.
        """
        start = str(self.image_dir if self.image_dir.exists() else Path.home())
        folder = QFileDialog.getExistingDirectory(
            self, "Select OCT data folder", start
        )
        if not folder:
            return
        folder = Path(folder)
        # The folder must contain at least one TIFF for the editor to be
        # useful. Glob silently to keep this method fast.
        tiffs = list(folder.glob("*.tif")) + list(folder.glob("*.tiff"))
        if not tiffs:
            QMessageBox.warning(
                self, "No TIFFs found",
                f"No .tif/.tiff files in {folder}.",
            )
            return
        workbook_path = folder / "output" / "oct_results.xlsx"
        if workbook_path.exists():
            self._switch_data_folder(folder, workbook_path)
            return
        # No analysis yet — offer to run it now.
        reply = QMessageBox.question(
            self, "No analysis found",
            f"No oct_results.xlsx found in:\n{workbook_path.parent}\n\n"
            f"Found {len(tiffs)} TIFF files in the folder.\n\n"
            "Run auto analysis now?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )
        if reply != QMessageBox.Yes:
            return
        # Switch the in-memory paths first so _run_auto_analysis uses them,
        # but skip _switch_data_folder's reload (the workbook does not
        # exist yet).
        self.image_dir = folder
        self.workbook_path = workbook_path
        self._editors.clear()
        self._current_stem = None
        self._run_auto_analysis()

    def _switch_data_folder(self, image_dir: Path,
                             workbook_path: Path) -> None:
        """Update paths and reload, prompting first if any editor is dirty."""
        dirty = [s for s, ed in self._editors.items() if ed.dirty]
        if dirty:
            n = len(dirty)
            msg = (f"Save changes to {n} image(s) before switching folder?"
                   if n > 1 else
                   f"Save changes to {self._wb.images[dirty[0]].filename} "
                   "before switching folder?")
            reply = QMessageBox.question(
                self, "Unsaved changes", msg,
                QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
                QMessageBox.Save,
            )
            if reply == QMessageBox.Cancel:
                return
            if reply == QMessageBox.Save:
                for stem in dirty:
                    self._current_stem = stem
                    self.save_current_image()
                # Wait for any background save to complete before switch.
                self._wait_for_save()
                if any(ed.dirty for ed in self._editors.values()):
                    return  # save failed — abort the switch
        self.image_dir = Path(image_dir)
        self.workbook_path = Path(workbook_path)
        self._editors.clear()
        self._current_stem = None
        # Discard cached openpyxl workbook so the new folder loads fresh.
        if self._openpyxl_wb is not None:
            try:
                self._openpyxl_wb.close()
            except Exception:
                pass
        self._openpyxl_wb = None
        self._reload_workbook()
        self.statusBar().showMessage(f"Loaded {self.workbook_path}")

    def _wait_for_save(self, timeout_ms: int = 60_000) -> bool:
        """Block until the background save thread (if any) completes.

        Returns True on success, False on timeout.
        """
        if self._save_thread is None:
            return True
        # Process events while waiting so the worker's finished signal
        # gets dispatched on the main thread.
        from PySide6.QtCore import QCoreApplication
        deadline = timeout_ms
        step = 50
        while self._save_in_flight and deadline > 0:
            QCoreApplication.processEvents()
            QThread.msleep(step)
            deadline -= step
        return not self._save_in_flight

    # ----------------------------------------------------- batch runner

    def _run_auto_analysis(self) -> None:
        """Run batch_process.batch_run on the current folder in a worker."""
        if not self.image_dir.exists():
            QMessageBox.critical(
                self, "No folder",
                "No data folder selected. Use File > Open Data Folder.",
            )
            return
        tiffs = list(self.image_dir.glob("*.tif")) + \
                list(self.image_dir.glob("*.tiff"))
        if not tiffs:
            QMessageBox.critical(
                self, "No TIFFs",
                f"No .tif/.tiff files in {self.image_dir}.",
            )
            return
        reply = QMessageBox.question(
            self, "Run auto analysis",
            f"Analyze {len(tiffs)} TIFF files in:\n{self.image_dir}\n\n"
            "This may take several minutes. Existing automatic results "
            "will be overwritten; *_corrected columns are preserved.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return
        if self._batch_thread is not None:
            QMessageBox.information(
                self, "Already running",
                "An auto analysis run is already in progress.",
            )
            return
        output_dir = self.image_dir / "output"
        # Spin up a worker on its own QThread.
        self._batch_thread = QThread(self)
        self._batch_worker = BatchWorker(self.image_dir, output_dir)
        self._batch_worker.moveToThread(self._batch_thread)
        self._batch_thread.started.connect(self._batch_worker.run)
        self._batch_worker.progress.connect(self._on_batch_progress)
        self._batch_worker.finished.connect(self._on_batch_finished)
        self._batch_worker.error.connect(self._on_batch_error)
        # Modal progress dialog. Cancel currently does not interrupt the
        # worker — it just hides the dialog; the run completes in the
        # background. (Adding interruption would need cooperative checks
        # in batch_run; out of scope.)
        self._progress_dialog = QProgressDialog(
            "Starting auto analysis...", None, 0, len(tiffs), self
        )
        self._progress_dialog.setWindowTitle("Auto Analysis")
        self._progress_dialog.setWindowModality(Qt.WindowModal)
        self._progress_dialog.setMinimumDuration(0)
        self._progress_dialog.setValue(0)
        self._batch_thread.start()

    def _on_batch_progress(self, i: int, n: int, name: str) -> None:
        if self._progress_dialog is None:
            return
        self._progress_dialog.setMaximum(n)
        self._progress_dialog.setValue(i)
        self._progress_dialog.setLabelText(f"[{i}/{n}] {name}")

    def _cleanup_batch_worker(self) -> None:
        if self._batch_thread is not None:
            self._batch_thread.quit()
            self._batch_thread.wait()
            self._batch_thread = None
        self._batch_worker = None
        if self._progress_dialog is not None:
            self._progress_dialog.close()
            self._progress_dialog = None

    def _on_batch_finished(self, xlsx_path: str) -> None:
        self._cleanup_batch_worker()
        self.workbook_path = Path(xlsx_path)
        # Editors and the cached openpyxl workbook are stale after
        # re-analysis (the on-disk xlsx changed); rebuild from scratch.
        self._editors.clear()
        self._current_stem = None
        if self._openpyxl_wb is not None:
            try:
                self._openpyxl_wb.close()
            except Exception:
                pass
        self._openpyxl_wb = None
        self._reload_workbook()
        QMessageBox.information(
            self, "Done",
            f"Auto analysis complete:\n{xlsx_path}",
        )

    def _on_batch_error(self, msg: str) -> None:
        self._cleanup_batch_worker()
        QMessageBox.critical(self, "Auto analysis failed", msg)

    def _prev_image(self) -> None:
        """Move sidebar selection up by one (clamped)."""
        count = self.sidebar.count()
        if count == 0:
            return
        row = self.sidebar.currentRow()
        new_row = max(0, row - 1)
        if new_row != row:
            self.sidebar.setCurrentRow(new_row)

    def _next_image(self) -> None:
        """Move sidebar selection down by one (clamped)."""
        count = self.sidebar.count()
        if count == 0:
            return
        row = self.sidebar.currentRow()
        new_row = min(count - 1, row + 1)
        if new_row != row:
            self.sidebar.setCurrentRow(new_row)

    def closeEvent(self, event) -> None:
        """Prompt save/discard/cancel if any image has unsaved edits."""
        dirty_stems = [stem for stem, ed in self._editors.items() if ed.dirty]
        if not dirty_stems:
            super().closeEvent(event)
            return
        n = len(dirty_stems)
        if n == 1 and self._wb is not None:
            msg = (f"Save changes to "
                   f"{self._wb.images[dirty_stems[0]].filename} "
                   f"before closing?")
        else:
            msg = (f"You have unsaved changes in {n} images. "
                   f"Save before closing?")
        reply = QMessageBox.question(
            self, "Unsaved changes", msg,
            QMessageBox.Save | QMessageBox.Discard | QMessageBox.Cancel,
            QMessageBox.Save,
        )
        if reply == QMessageBox.Cancel:
            event.ignore()
            return
        if reply == QMessageBox.Save:
            # Queue saves for each dirty image, then block until the
            # background worker finishes flushing.
            for stem in dirty_stems:
                self._current_stem = stem
                try:
                    self.save_current_image()
                except Exception:
                    event.ignore()
                    return
            # save_current_image is async — wait for the worker.
            self._wait_for_save()
            if any(ed.dirty for ed in self._editors.values()):
                event.ignore()
                return
        else:  # Discard
            # Even on discard, wait for any in-flight save so we don't
            # tear down the workbook mid-write.
            self._wait_for_save()
        super().closeEvent(event)

    def _on_canvas_hover(self, x_local: int) -> None:
        """Update the rightmost status-bar label with per-column readings."""
        if x_local < 0 or self._current_stem is None:
            self._hover_label.setText("")
            return
        editor = self._editors.get(self._current_stem)
        if editor is None:
            self._hover_label.setText("")
            return

        def fmt(v) -> str:
            if v is None or (isinstance(v, float) and np.isnan(v)):
                return "—"
            return f"{int(round(float(v)))}"

        eff = {n: editor.effective(n)[x_local] for n in BOUNDARY_NAMES}
        parts = [f"x={x_local}",
                 f"TOP={fmt(eff['TOP_y'])}",
                 f"ONL={fmt(eff['ONL_y'])}",
                 f"BM={fmt(eff['BM_y'])}"]
        # DET pair only when at least the top is finite.
        det_top = eff["DET_top_y"]
        det_bot = eff["DET_bottom_y"]
        if not (isinstance(det_top, float) and np.isnan(det_top)):
            parts.append(f"DET=({fmt(det_top)},{fmt(det_bot)})")
        # Thicknesses (μm) — only when both endpoints are finite.
        scale = self._scale_y
        bm = eff["BM_y"]
        top = eff["TOP_y"]
        onl = eff["ONL_y"]

        def finite(v) -> bool:
            return not (isinstance(v, float) and np.isnan(v))

        if finite(bm) and finite(top):
            parts.append(f"total={(bm - top) * scale:.1f}μm")
        if finite(bm) and finite(onl):
            parts.append(f"outer={(bm - onl) * scale:.1f}μm")
        if finite(det_top) and finite(det_bot):
            parts.append(f"det={(det_bot - det_top) * scale:.1f}μm")
        self._hover_label.setText("  |  ".join(parts))

    def _refresh_status(self) -> None:
        """Rebuild the status bar message with edited boundaries + dirty marker."""
        if self._current_stem is None or self._wb is None:
            return
        if self._current_stem not in self._wb.images:
            return
        rec = self._wb.images[self._current_stem]
        editor = self._editors.get(self._current_stem)
        if editor is None:
            return
        edited: list[str] = []
        for name in BOUNDARY_NAMES:
            if name not in editor.corrected:
                continue
            corr = editor.corrected[name]
            # "touched" = any column either set to a finite override or
            # explicitly erased (sentinel below threshold).
            touched = bool(np.any((~np.isnan(corr)) | (corr < ERASED_THRESHOLD)))
            if touched:
                edited.append(name.replace("_y", ""))
        parts = [rec.filename, f"{rec.width} cols"]
        if edited:
            parts.append("edited: " + ", ".join(edited))
        if editor.dirty:
            parts.append("●")
        self.statusBar().showMessage(" | ".join(parts))


def run(workbook_path, image_dir) -> None:
    app = QApplication.instance() or QApplication(sys.argv)
    win = MainWindow(workbook_path=workbook_path, image_dir=image_dir)
    win.resize(1400, 800)
    win.show()
    app.exec()
